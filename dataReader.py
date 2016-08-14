from openpyxl import Workbook
import openpyxl
import json
import subprocess

workbook = openpyxl.load_workbook('TextEx.xlsx')
workbook.save('new_ex.xlsx')

worksheet = workbook.active

# Get iter rows
iter_rows = worksheet.iter_rows()

# Add to row list
row_list = []
for row in iter_rows:
    row_list.append(row)

# Get header list
header_list = []
for header_row in row_list[0]:
    # print(header_row.value)
    header_list.append(header_row.value.lower())

row_list.pop(0)

# Prepare data list
data_list = []

for row in row_list:
    cell_dict = {}
    for index, cell in enumerate(row):
        cell_dict[header_list[index]] = cell.value

    data_list.append(cell_dict)

# Convert to JSON
data_json = json.dumps(data_list)

# Write file
with open('data.js', 'w') as f:
    f.write('var data = ')
    f.write(data_json)

subprocess.check_output('start games.html', shell=True)
